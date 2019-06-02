
import uuid

import openpyxl

import dbConnector

import os

import config

db = dbConnector.initDB();

class student:
        def __init__(self, user_id, first_name, last_name, email, program, level):
                self.user_id = user_id
                self.first_name = first_name
                self.last_name = last_name
                self.email = email
                self.program = program
                self.level = level

        def process_currentaccounts(foundAccounts):

                 currentAccounts= []

                 for account in foundAccounts:

                        currentAccounts.append(account);

                 return currentAccounts;

        def process_roster(currentAccounts, path):

                roster= []

                # workbook object is created
                wb_obj = openpyxl.load_workbook(path)

                sheet_obj = wb_obj.active
                col_count = sheet_obj.max_column
                row_count = sheet_obj.max_row

                # Loop will print all columns name
                for row in sheet_obj.iter_rows(min_row=2, max_col=col_count, max_row=row_count, values_only=True):


                        email = row[3];
                        thisName = row[4].split(",", 2);
                        last_name = thisName[0];
                        first_name = thisName[1];
                        program = row[7];
                        level = row[8];
                        id = str(uuid.uuid4());

                        thisstudent = student(id, first_name, last_name, email, program, level)
                        roster.append(thisstudent);

                return roster;


        def accountCheck(currentAccounts, roster):

                for indiv in roster:

                        match = False;

                        for user in currentAccounts:

                             if user.email == indiv.email:
                                                match = True;


                        if match == False:

                                student.addUserToDB(indiv)
                                print('Added student' + indiv.email + ' to DB')
                        else:
                                 print(indiv.email + 'Already Existed in DB')

        def addUserToDB(newUser):

                cursor = db.cursor()

                sql_insert_query = ( "INSERT INTO user"
                                     "(user_id, first_name, last_name, email, program, level) "
                                     "VALUES (%s,%s,%s,%s,%s,%s)")


                insertThis = (newUser.user_id, newUser.first_name, newUser.last_name, newUser.email, newUser.program, newUser.level)

                result = cursor.execute(sql_insert_query, insertThis)

                db.commit()

                group_id = config.course_id;

                member_id = newUser.user_id;

                student.add_group_member(group_id, member_id)


        def add_group_member(group_id, member_id):

                cursor = db.cursor()

                sql_insert_query = ( "INSERT INTO ce301.grouping_member "
                                     "(group_id, related_id) "
                                     "VALUES (%s,%s)")


                insertThis = (group_id, member_id)

                result = cursor.execute(sql_insert_query, insertThis)

                db.commit()
                print ('Record' + member_id + 'inserted successfully into grouping member')

        def drop_group_member(group_id, member_id):

                cursor = db.cursor()

                sql_insert_query = ( "Delete from ce301.grouping_member "
                                     "where group_id = %s and related_id = %s ")

                deletetThis = (group_id, member_id)

                result = cursor.execute(sql_insert_query, deletetThis)

                db.commit()
                print ('Record' + member_id + ' successfully deleted')




        def post_processing_check(roster):


                print("Post Check")
                postcurso = db.cursor(named_tuple=True)

                sql_insert_query = ("select u.* from user u "
                                    "join grouping_member gm on gm.related_id = u.user_id "
                                    "where level is not null "
                                    "and group_id = %s ")


                # getting records from the table
                postcurso.execute(sql_insert_query, (config.course_id,))

                # fetching all records from the 'cursor' object
                accounts = postcurso.fetchall()

                for indiv in accounts:


                        match = False;

                        for user in roster:

                                if user.email == indiv.email:
                                        match = True;


                        if match == False:

                                student.drop_group_member(course_id, indiv.user_id)
                                print('Dropped student' + indiv.email + ' to DB')
                        else:
                                pass




cursor = db.cursor(named_tuple=True)

query = "select * from user where level is not null "

# getting records from the table
cursor.execute(query)

# fetching all records from the 'cursor' object
records = cursor.fetchall()

print("Fetched Records")

rosterPath = os.getcwd() + "/roster.xlsx";

currentAccounts = student.process_currentaccounts(records);

roster = student.process_roster(currentAccounts, rosterPath);

crossCompare = student.accountCheck(currentAccounts, roster);

dropCheck = student.post_processing_check(roster);



